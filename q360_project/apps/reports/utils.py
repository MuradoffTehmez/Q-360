"""
Utility functions for report generation.
"""
from io import BytesIO
from django.conf import settings
from apps.evaluations.models import Response, QuestionCategory


def calculate_radar_data(evaluation_result):
    """Calculate radar chart data for an evaluation result."""
    from apps.evaluations.models import EvaluationAssignment

    assignments = EvaluationAssignment.objects.filter(
        campaign=evaluation_result.campaign,
        evaluatee=evaluation_result.evaluatee,
        status='completed'
    )

    categories = QuestionCategory.objects.filter(is_active=True)

    radar_data = {
        'categories': [],
        'self': [],
        'others': [],
        'average': []
    }

    for category in categories:
        radar_data['categories'].append(category.name)

        # Scores by relationship
        self_scores = []
        other_scores = []

        for assignment in assignments:
            responses = Response.objects.filter(
                assignment=assignment,
                question__category=category,
                score__isnull=False
            )

            for response in responses:
                if assignment.relationship == 'self':
                    self_scores.append(response.score)
                else:
                    other_scores.append(response.score)

        # Calculate averages
        self_avg = sum(self_scores) / len(self_scores) if self_scores else 0
        other_avg = sum(other_scores) / len(other_scores) if other_scores else 0
        overall_avg = (self_avg + other_avg) / 2 if (self_avg or other_avg) else 0

        radar_data['self'].append(round(self_avg, 2))
        radar_data['others'].append(round(other_avg, 2))
        radar_data['average'].append(round(overall_avg, 2))

    return radar_data


def generate_pdf_report(evaluation_result):
    """Generate PDF report for evaluation result."""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        # Fallback if reportlab not installed
        return b"PDF generation requires reportlab package"

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []

    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#0d6efd'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    story.append(Paragraph('360° Qiymətləndirmə Hesabatı', title_style))
    story.append(Spacer(1, 20))

    # Employee info
    info_data = [
        ['İşçi:', evaluation_result.evaluatee.get_full_name()],
        ['Kampaniya:', evaluation_result.campaign.title],
        ['Tarix:', evaluation_result.calculated_at.strftime('%d.%m.%Y')],
        ['Ümumi Bal:', f'{evaluation_result.overall_score:.2f}' if evaluation_result.overall_score else 'N/A'],
    ]

    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    story.append(info_table)
    story.append(Spacer(1, 30))

    # Scores by relationship
    scores_data = [
        ['Qiymətləndirmə Növü', 'Ortalama Bal'],
        ['Özünüdəyərləndirmə', f'{evaluation_result.self_score:.2f}' if evaluation_result.self_score else '-'],
        ['Rəhbər', f'{evaluation_result.supervisor_score:.2f}' if evaluation_result.supervisor_score else '-'],
        ['Həmkar', f'{evaluation_result.peer_score:.2f}' if evaluation_result.peer_score else '-'],
        ['Tabelik', f'{evaluation_result.subordinate_score:.2f}' if evaluation_result.subordinate_score else '-'],
    ]

    scores_table = Table(scores_data, colWidths=[3*inch, 3*inch])
    scores_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d6efd')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    story.append(Paragraph('Qiymətləndirmə Nəticələri', styles['Heading2']))
    story.append(Spacer(1, 12))
    story.append(scores_table)

    # Build PDF
    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()

    return pdf_content


def generate_excel_report(campaign):
    """Generate Excel report for campaign results."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return b"Excel generation requires openpyxl package"

    from apps.evaluations.models import EvaluationResult

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nəticələr"

    # Header style
    header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    # Headers
    headers = [
        'İşçi ID', 'Ad Soyad', 'Şöbə', 'Vəzifə',
        'Ümumi Bal', 'Özüm', 'Rəhbər', 'Həmkar', 'Tabelik',
        'Qiymətləndirən Sayı', 'Tamamlanma %'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    results = EvaluationResult.objects.filter(
        campaign=campaign
    ).select_related('evaluatee', 'evaluatee__department').order_by('-overall_score')

    for row, result in enumerate(results, start=2):
        ws.cell(row=row, column=1).value = result.evaluatee.employee_id or result.evaluatee.username
        ws.cell(row=row, column=2).value = result.evaluatee.get_full_name()
        ws.cell(row=row, column=3).value = str(result.evaluatee.department) if result.evaluatee.department else '-'
        ws.cell(row=row, column=4).value = result.evaluatee.position or '-'
        ws.cell(row=row, column=5).value = float(result.overall_score) if result.overall_score else 0
        ws.cell(row=row, column=6).value = float(result.self_score) if result.self_score else 0
        ws.cell(row=row, column=7).value = float(result.supervisor_score) if result.supervisor_score else 0
        ws.cell(row=row, column=8).value = float(result.peer_score) if result.peer_score else 0
        ws.cell(row=row, column=9).value = float(result.subordinate_score) if result.subordinate_score else 0
        ws.cell(row=row, column=10).value = result.total_evaluators
        ws.cell(row=row, column=11).value = float(result.completion_rate)

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    excel_content = buffer.getvalue()
    buffer.close()

    return excel_content
