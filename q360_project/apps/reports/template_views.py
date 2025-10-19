"""
Template views for reports app.
Professional report generation and visualization.
"""
from datetime import timedelta
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Avg, Count, Q, Max, Min
from django.utils import timezone
import json

from apps.evaluations.models import EvaluationResult, EvaluationCampaign, Response, EvaluationAssignment
from apps.accounts.models import User
from apps.accounts.permissions import filter_queryset_for_user, get_accessible_users
from .models import (
    Report,
    RadarChartData,
    ReportBlueprint,
    ReportSchedule,
    ReportScheduleLog,
    ReportVisualization,
    SystemKPI,
)
from .utils import generate_pdf_report, generate_excel_report, generate_csv_report, calculate_radar_data

from .services import build_dataset_for_blueprint

@login_required
def my_reports(request):
    """View current user's evaluation reports."""
    user = request.user

    # Get user's results
    results = EvaluationResult.objects.filter(
        evaluatee=user
    ).select_related('campaign').order_by('-calculated_at')

    # Latest result
    latest_result = results.first()

    # Performance trend data
    trend_data = {
        'labels': [],
        'data': []
    }

    for result in results[:6]:  # Last 6 evaluations
        trend_data['labels'].insert(0, result.campaign.title[:20])
        trend_data['data'].insert(0, float(result.overall_score) if result.overall_score else 0)

    # Radar chart data for latest result
    radar_data = None
    if latest_result:
        radar_data = calculate_radar_data(latest_result)

    context = {
        'results': results,
        'latest_result': latest_result,
        'trend_data': json.dumps(trend_data),
        'radar_data': json.dumps(radar_data) if radar_data else None,
    }

    return render(request, 'reports/my_reports.html', context)


@login_required
def team_reports(request):
    """View team evaluation reports (for managers)."""
    if not request.user.is_manager():
        messages.error(request, 'Bu səhifəyə giriş icazəniz yoxdur.')
        return redirect('dashboard')

    user = request.user

    # Get subordinates
    team_members = get_accessible_users(user)
    if not user.is_admin():
        team_members = team_members.exclude(pk=user.pk)

    # Get campaign from query parameter or latest
    campaign_id = request.GET.get('campaign')
    if campaign_id:
        latest_campaign = EvaluationCampaign.objects.filter(
            pk=campaign_id,
            status__in=['active', 'completed']
        ).first()
    else:
        latest_campaign = EvaluationCampaign.objects.filter(
            status__in=['active', 'completed']
        ).order_by('-created_at').first()

    # Get results for team members
    results = None
    if latest_campaign:
        result_qs = EvaluationResult.objects.filter(
            campaign=latest_campaign
        ).select_related('evaluatee').order_by('-overall_score')
        results = filter_queryset_for_user(
            user,
            result_qs,
            relation_field='evaluatee'
        )

    # Calculate team statistics
    team_stats = {
        'total_members': team_members.count() if hasattr(team_members, 'count') else len(team_members),
        'evaluated_members': results.count() if results else 0,
        'avg_score': results.aggregate(Avg('overall_score'))['overall_score__avg'] if results else None,
        'top_performers': list(results[:5]) if results else [],
        'needs_attention': list(results.order_by('overall_score')[:5]) if results else [],
    }

    # Department comparison data
    dept_comparison = []
    if user.is_admin() and results:
        from apps.departments.models import Department
        departments = Department.objects.filter(is_active=True)

        for dept in departments:
            dept_results = results.filter(evaluatee__department=dept)
            if dept_results.exists():
                dept_comparison.append({
                    'name': dept.name,
                    'avg_score': dept_results.aggregate(Avg('overall_score'))['overall_score__avg'],
                    'count': dept_results.count()
                })

    context = {
        'team_members': team_members,
        'results': results,
        'latest_campaign': latest_campaign,
        'team_stats': team_stats,
        'dept_comparison': dept_comparison,
    }

    return render(request, 'reports/team_reports.html', context)


@login_required
def blueprint_list(request):
    """Blueprint listing for custom report definitions."""
    blueprints = ReportBlueprint.objects.filter(is_active=True).prefetch_related(
        'visualizations',
        'schedules',
        'owner',
    ).order_by('title')

    source = request.GET.get('source')
    export = request.GET.get('export')
    if source:
        blueprints = blueprints.filter(data_source=source)
    if export:
        blueprints = blueprints.filter(default_export_format=export)
    if request.GET.get('global'):
        blueprints = blueprints.filter(is_global=True)

    context = {
        'blueprints': blueprints,
        'sources': ReportBlueprint.DATA_SOURCE_CHOICES,
        'export_formats': ReportBlueprint.EXPORT_FORMAT_CHOICES,
    }
    return render(request, 'reports/blueprint_list.html', context)


@login_required
def blueprint_detail(request, slug):
    """Detailed view for a single blueprint."""
    blueprint = get_object_or_404(
        ReportBlueprint.objects.prefetch_related('visualizations', 'schedules__recipients'),
        slug=slug,
    )
    dataset = build_dataset_for_blueprint(blueprint)
    preview = {
        'headers': dataset.columns,
        'rows': dataset.rows[:5],
    }
    schedules = list(
        blueprint.schedules.select_related('created_by').prefetch_related('recipients').order_by('next_run')
    )
    status_badges = {
        'completed': 'success',
        'failed': 'danger',
        'processing': 'info',
        'pending': 'secondary',
    }
    for schedule in schedules:
        schedule.last_status_badge = status_badges.get(schedule.last_status, 'secondary')

    context = {
        'blueprint': blueprint,
        'visualizations': blueprint.visualizations.all().order_by('order'),
        'column_config': blueprint.columns or [],
        'schedules': schedules,
        'preview': preview,
        'dataset_metadata': dataset.metadata,
    }
    return render(request, 'reports/blueprint_detail.html', context)


@login_required
def schedule_center(request):
    """Overview for scheduled report exports."""
    schedules = ReportSchedule.objects.select_related('blueprint').prefetch_related('recipients').order_by('-next_run')
    status_badges = {
        'completed': 'success',
        'failed': 'danger',
        'processing': 'info',
        'pending': 'secondary',
    }
    for schedule in schedules:
        schedule.last_status_badge = status_badges.get(schedule.last_status, 'secondary')

    now = timezone.now()
    stats = {
        'active': schedules.filter(is_active=True).count(),
        'paused': schedules.filter(is_active=False).count(),
        'failed': schedules.filter(last_status='failed').count(),
        'next_run': schedules.filter(is_active=True, next_run__isnull=False).order_by('next_run').values_list('next_run', flat=True).first(),
        'sent_last_7_days': ReportScheduleLog.objects.filter(
            status='completed',
            triggered_at__gte=now - timedelta(days=7)
        ).count(),
    }

    recent_logs = list(
        ReportScheduleLog.objects.select_related('schedule', 'schedule__blueprint', 'export_log')
        .order_by('-triggered_at')[:6]
    )

    context = {
        'schedules': schedules,
        'stats': stats,
        'recent_logs': recent_logs,
    }
    return render(request, 'reports/schedule_center.html', context)


@login_required
def detailed_report(request, result_pk):
    """View detailed individual report."""
    result = get_object_or_404(
        EvaluationResult.objects.select_related('campaign', 'evaluatee'),
        pk=result_pk
    )

    # Check permission
    if not (request.user.is_admin() or
            result.evaluatee == request.user or
            result.evaluatee.supervisor == request.user):
        messages.error(request, 'Bu hesabata baxmaq icazəniz yoxdur.')
        return redirect('dashboard')

    # Get all assignments for this result
    assignments = EvaluationAssignment.objects.filter(
        campaign=result.campaign,
        evaluatee=result.evaluatee,
        status='completed'
    ).select_related('evaluator')

    # Calculate category-wise scores
    from apps.evaluations.models import QuestionCategory
    categories = QuestionCategory.objects.filter(is_active=True)

    category_analysis = []
    for category in categories:
        scores_by_relationship = {
            'self': [],
            'supervisor': [],
            'peer': [],
            'subordinate': []
        }

        for assignment in assignments:
            responses = Response.objects.filter(
                assignment=assignment,
                question__category=category,
                score__isnull=False
            )

            for response in responses:
                scores_by_relationship[assignment.relationship].append(response.score)

        # Calculate averages
        category_data = {
            'name': category.name,
            'scores': {}
        }

        for rel_type, scores in scores_by_relationship.items():
            if scores:
                category_data['scores'][rel_type] = round(sum(scores) / len(scores), 2)
            else:
                category_data['scores'][rel_type] = None

        # Overall average for category
        all_scores = []
        for scores in scores_by_relationship.values():
            all_scores.extend(scores)
        if all_scores:
            category_data['average'] = round(sum(all_scores) / len(all_scores), 2)
        else:
            category_data['average'] = None

        category_analysis.append(category_data)

    # Get text responses (comments)
    text_responses = Response.objects.filter(
        assignment__in=assignments,
        question__question_type='text'
    ).select_related('question').exclude(text_answer='')

    # Strengths and development areas
    strengths = text_responses.filter(
        Q(question__text__icontains='güclü') | Q(question__text__icontains='strength')
    )
    development = text_responses.filter(
        Q(question__text__icontains='inkişaf') | Q(question__text__icontains='development')
    )

    # Radar chart data
    radar_data = calculate_radar_data(result)

    context = {
        'result': result,
        'assignments': assignments,
        'category_analysis': category_analysis,
        'text_responses': text_responses,
        'strengths': strengths,
        'development': development,
        'radar_data': json.dumps(radar_data),
    }

    return render(request, 'reports/detailed_report.html', context)


@login_required
def export_pdf(request, result_pk):
    """Export report as PDF."""
    result = get_object_or_404(EvaluationResult, pk=result_pk)

    # Check permission
    if not (request.user.is_admin() or result.evaluatee == request.user):
        messages.error(request, 'Bu hesabatı ixrac etmək icazəniz yoxdur.')
        return redirect('dashboard')

    # Generate PDF
    pdf_content = generate_pdf_report(result)

    response = HttpResponse(pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="qiymetlendirme_{result.evaluatee.username}_{result.campaign.pk}.pdf"'

    return response


@login_required
def export_excel(request, campaign_pk):
    """Export campaign results as Excel."""
    campaign = get_object_or_404(EvaluationCampaign, pk=campaign_pk)

    # Check permission
    if not request.user.is_admin():
        messages.error(request, 'Bu hesabatı ixrac etmək icazəniz yoxdur.')
        return redirect('dashboard')

    # Generate Excel
    excel_content = generate_excel_report(campaign)

    response = HttpResponse(
        excel_content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="kampaniya_{campaign.pk}_neticeler.xlsx"'

    return response


@login_required
def comparison_report(request):
    """Compare multiple evaluation results."""
    user = request.user

    # Get selected campaigns or results
    campaign_ids = request.GET.getlist('campaigns')

    if not campaign_ids:
        # Show selection form
        campaigns = EvaluationCampaign.objects.filter(
            status__in=['completed', 'active']
        ).order_by('-created_at')[:10]

        context = {
            'campaigns': campaigns
        }
        return render(request, 'reports/comparison_select.html', context)

    # Get results for comparison
    results = EvaluationResult.objects.filter(
        campaign_id__in=campaign_ids,
        evaluatee=user
    ).select_related('campaign').order_by('campaign__start_date')

    # Prepare comparison data
    comparison_data = {
        'labels': [],
        'overall': [],
        'self': [],
        'supervisor': [],
        'peer': [],
        'subordinate': []
    }

    for result in results:
        comparison_data['labels'].append(result.campaign.title[:20])
        comparison_data['overall'].append(float(result.overall_score) if result.overall_score else 0)
        comparison_data['self'].append(float(result.self_score) if result.self_score else 0)
        comparison_data['supervisor'].append(float(result.supervisor_score) if result.supervisor_score else 0)
        comparison_data['peer'].append(float(result.peer_score) if result.peer_score else 0)
        comparison_data['subordinate'].append(float(result.subordinate_score) if result.subordinate_score else 0)

    context = {
        'results': results,
        'comparison_data': json.dumps(comparison_data),
    }

    return render(request, 'reports/comparison_report.html', context)


@login_required
def analytics_dashboard(request):
    """
    Advanced analytics dashboard with comprehensive KPIs (admin only).
    Displays system-wide metrics, trends, and insights.
    """
    if not request.user.is_admin():
        messages.error(request, 'Bu səhifəyə giriş icazəniz yoxdur.')
        return redirect('dashboard')

    from datetime import date
    from apps.departments.models import Department
    from apps.training.models import TrainingResource
    from apps.audit.models import AuditLog

    # Get or calculate today's KPI
    today = date.today()
    today_kpi = SystemKPI.objects.filter(date=today).first()

    if not today_kpi:
        # Calculate if doesn't exist
        today_kpi = SystemKPI.calculate_today_kpis()

    # Get KPI trend data (last 30 days)
    thirty_days_ago = today - timedelta(days=30)
    kpi_history = SystemKPI.objects.filter(
        date__gte=thirty_days_ago
    ).order_by('date')

    # Prepare trend data for charts
    user_trend = {
        'labels': [],
        'active_users': [],
        'new_users': []
    }

    evaluation_trend = {
        'labels': [],
        'completion_rate': [],
        'completed_count': []
    }

    security_trend = {
        'labels': [],
        'login_attempts': [],
        'failed_attempts': []
    }

    for kpi in kpi_history:
        date_label = kpi.date.strftime('%d %b')

        # User trends
        user_trend['labels'].append(date_label)
        user_trend['active_users'].append(kpi.active_users)
        user_trend['new_users'].append(kpi.new_users_today)

        # Evaluation trends
        evaluation_trend['labels'].append(date_label)
        evaluation_trend['completion_rate'].append(float(kpi.completion_rate))
        evaluation_trend['completed_count'].append(kpi.evaluations_completed_today)

        # Security trends
        security_trend['labels'].append(date_label)
        security_trend['login_attempts'].append(kpi.login_attempts_today)
        security_trend['failed_attempts'].append(kpi.failed_login_attempts_today)

    # Department performance comparison
    departments = Department.objects.filter(is_active=True)
    dept_performance = []

    latest_campaign = EvaluationCampaign.objects.filter(
        status__in=['active', 'completed']
    ).order_by('-created_at').first()

    if latest_campaign:
        for dept in departments:
            dept_results = EvaluationResult.objects.filter(
                campaign=latest_campaign,
                evaluatee__department=dept
            )

            if dept_results.exists():
                avg_score = dept_results.aggregate(Avg('overall_score'))['overall_score__avg']
                dept_performance.append({
                    'name': dept.name,
                    'avg_score': round(float(avg_score), 2) if avg_score else 0,
                    'count': dept_results.count(),
                    'employees': User.objects.filter(department=dept, is_active=True).count()
                })

    # Sort by average score
    dept_performance.sort(key=lambda x: x['avg_score'], reverse=True)

    # Top performers (last campaign)
    top_performers = []
    if latest_campaign:
        top_results = EvaluationResult.objects.filter(
            campaign=latest_campaign
        ).select_related('evaluatee').order_by('-overall_score')[:10]

        for result in top_results:
            top_performers.append({
                'name': result.evaluatee.get_full_name(),
                'department': str(result.evaluatee.department) if result.evaluatee.department else '-',
                'position': result.evaluatee.position or '-',
                'score': float(result.overall_score) if result.overall_score else 0
            })

    # System health metrics
    system_health = {
        'database_size': float(today_kpi.database_size_mb) if today_kpi.database_size_mb else 0,
        'response_time': float(today_kpi.average_response_time) if today_kpi.average_response_time else 0,
        'active_users_percent': round((today_kpi.active_users / today_kpi.total_users * 100) if today_kpi.total_users > 0 else 0, 1),
        'login_success_rate': round(
            ((today_kpi.login_attempts_today - today_kpi.failed_login_attempts_today) / today_kpi.login_attempts_today * 100)
            if today_kpi.login_attempts_today > 0 else 100, 1
        )
    }

    # Recent activity summary
    recent_activities = []

    # New users today
    new_users_today = User.objects.filter(date_joined__date=today).select_related('department')[:5]
    for user in new_users_today:
        recent_activities.append({
            'type': 'user_joined',
            'icon': 'fa-user-plus',
            'color': 'blue',
            'message': f'{user.get_full_name()} qeydiyyatdan keçdi',
            'time': user.date_joined
        })

    # Recent evaluations
    recent_evals = EvaluationAssignment.objects.filter(
        completed_at__date=today
    ).select_related('evaluatee', 'evaluator')[:5]
    for eval in recent_evals:
        recent_activities.append({
            'type': 'evaluation_completed',
            'icon': 'fa-check-circle',
            'color': 'green',
            'message': f'{eval.evaluator.get_full_name()} tərəfindən {eval.evaluatee.get_full_name()} qiymətləndirildi',
            'time': eval.completed_at
        })

    # Sort by time
    recent_activities.sort(key=lambda x: x['time'], reverse=True)
    recent_activities = recent_activities[:10]

    # Score distribution for latest campaign
    score_distribution = {
        'labels': ['0-2', '2-3', '3-4', '4-5'],
        'data': [0, 0, 0, 0]
    }

    if latest_campaign:
        results = EvaluationResult.objects.filter(campaign=latest_campaign)
        for result in results:
            if result.overall_score:
                score = float(result.overall_score)
                if score < 2:
                    score_distribution['data'][0] += 1
                elif score < 3:
                    score_distribution['data'][1] += 1
                elif score < 4:
                    score_distribution['data'][2] += 1
                else:
                    score_distribution['data'][3] += 1

    context = {
        # Today's KPIs
        'today_kpi': today_kpi,

        # Trend data (JSON for charts)
        'user_trend': json.dumps(user_trend),
        'evaluation_trend': json.dumps(evaluation_trend),
        'security_trend': json.dumps(security_trend),

        # Department comparison
        'dept_performance': dept_performance,

        # Top performers
        'top_performers': top_performers,

        # System health
        'system_health': system_health,

        # Recent activities
        'recent_activities': recent_activities,

        # Score distribution
        'score_distribution': json.dumps(score_distribution),

        # Campaign info
        'latest_campaign': latest_campaign,
    }

    return render(request, 'reports/analytics_dashboard.html', context)


@login_required
def custom_report_builder(request):
    """
    Custom Report Builder - İstifadəçilərə özlərinə lazım olan hesabatı yaratmaq imkanı.

    Seçimlər:
    - Kampaniya seçimi
    - İstifadəçi/Şöbə filtri
    - Göstəriləcək metriklər
    - Chart növü
    """
    if request.method == 'POST':
        # Process custom report generation
        campaign_id = request.POST.get('campaign')
        department_id = request.POST.get('department')
        user_ids = request.POST.getlist('users')

        # Metrics to include
        include_overall = request.POST.get('include_overall') == 'on'
        include_category = request.POST.get('include_category') == 'on'
        include_relationship = request.POST.get('include_relationship') == 'on'
        include_trends = request.POST.get('include_trends') == 'on'

        # Chart types
        chart_types = request.POST.getlist('chart_types')

        # Build query
        results_query = EvaluationResult.objects.select_related('evaluatee', 'campaign')

        if campaign_id:
            results_query = results_query.filter(campaign_id=campaign_id)

        if department_id:
            results_query = results_query.filter(evaluatee__department_id=department_id)

        if user_ids:
            results_query = results_query.filter(evaluatee_id__in=user_ids)

        results = results_query.all()

        # Prepare data structures
        report_data = {
            'results': results,
            'summary': {},
            'charts': {}
        }

        # Overall statistics
        if include_overall:
            report_data['summary']['overall'] = {
                'total_evaluations': results.count(),
                'avg_score': results.aggregate(Avg('overall_score'))['overall_score__avg'],
                'max_score': results.aggregate(max_score=Max('overall_score'))['max_score'],
                'min_score': results.aggregate(min_score=Min('overall_score'))['min_score'],
            }

        # Category analysis
        if include_category:
            from apps.evaluations.models import QuestionCategory
            categories = QuestionCategory.objects.filter(is_active=True)
            category_data = {'labels': [], 'data': []}

            for category in categories:
                # Get all responses for this category
                avg_score = Response.objects.filter(
                    assignment__campaign_id=campaign_id,
                    question__category=category,
                    score__isnull=False
                ).aggregate(Avg('score'))['score__avg']

                if avg_score:
                    category_data['labels'].append(category.name)
                    category_data['data'].append(round(avg_score, 2))

            report_data['charts']['category'] = category_data

        # Relationship type analysis
        if include_relationship:
            relationship_data = {'labels': [], 'data': []}
            relationships = ['self', 'supervisor', 'peer', 'subordinate']

            for rel in relationships:
                if campaign_id:
                    avg_score = Response.objects.filter(
                        assignment__campaign_id=campaign_id,
                        assignment__relationship=rel,
                        score__isnull=False
                    ).aggregate(Avg('score'))['score__avg']

                    if avg_score:
                        relationship_data['labels'].append(rel.capitalize())
                        relationship_data['data'].append(round(avg_score, 2))

            report_data['charts']['relationship'] = relationship_data

        # Trend analysis
        if include_trends:
            trend_data = {'labels': [], 'data': []}
            # Get last 6 evaluations for selected users
            for result in results.order_by('campaign__start_date')[:6]:
                trend_data['labels'].append(result.campaign.title[:15])
                trend_data['data'].append(float(result.overall_score or 0))

            report_data['charts']['trend'] = trend_data

        # Convert to JSON for charts
        report_json = {
            'summary': report_data['summary'],
            'charts': {k: json.dumps(v) for k, v in report_data['charts'].items()}
        }

        context = {
            'report_data': report_data,
            'report_json': report_json,
            'chart_types': chart_types,
            'campaign_id': campaign_id,
        }

        return render(request, 'reports/custom_report_view.html', context)

    # GET: Show report builder form
    from apps.departments.models import Department

    campaigns = EvaluationCampaign.objects.filter(
        status__in=['completed', 'active']
    ).order_by('-created_at')

    departments = Department.objects.filter(is_active=True)

    # Get users based on permission
    if request.user.is_admin():
        users = User.objects.filter(is_active=True)
    elif request.user.is_manager():
        users = request.user.get_subordinates()
    else:
        users = [request.user]

    context = {
        'campaigns': campaigns,
        'departments': departments,
        'users': users,
    }

    return render(request, 'reports/custom_report_builder.html', context)


@login_required
def export_custom_report(request):
    """Export custom report in various formats (PDF/Excel/CSV)."""
    if request.method != 'POST':
        return redirect('reports:custom-builder')

    export_format = request.POST.get('export_format', 'excel')
    campaign_id = request.POST.get('campaign')
    department_id = request.POST.get('department')
    user_ids = request.POST.getlist('users')

    # Build query (same as custom_report_builder)
    results_query = EvaluationResult.objects.select_related('evaluatee', 'campaign', 'evaluatee__department')

    if campaign_id:
        results_query = results_query.filter(campaign_id=campaign_id)

    if department_id:
        results_query = results_query.filter(evaluatee__department_id=department_id)

    if user_ids:
        results_query = results_query.filter(evaluatee_id__in=user_ids)

    results = results_query.all()

    if not results.exists():
        messages.error(request, 'Seçilmiş meyarlara uyğun heç bir nəticə tapılmadı.')
        return redirect('reports:custom-builder')

    # Generate export based on format
    if export_format == 'pdf':
        # For PDF, we'll export first result as sample
        # In production, you might want to create a combined PDF
        if results.first():
            pdf_content = generate_pdf_report(results.first())
            response = HttpResponse(pdf_content, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="custom_report_{campaign_id}.pdf"'
            return response

    elif export_format == 'excel':
        # Generate Excel with all results
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Custom Report"

        # Header style
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)

        # Headers
        headers = [
            'İşçi ID', 'Ad Soyad', 'Şöbə', 'Vəzifə', 'Kampaniya',
            'Ümumi Bal', 'Özüm', 'Rəhbər', 'Həmkar', 'Tabelik', 'Tarix'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data
        for row, result in enumerate(results, start=2):
            ws.cell(row=row, column=1).value = result.evaluatee.employee_id or result.evaluatee.username
            ws.cell(row=row, column=2).value = result.evaluatee.get_full_name()
            ws.cell(row=row, column=3).value = str(result.evaluatee.department) if result.evaluatee.department else '-'
            ws.cell(row=row, column=4).value = result.evaluatee.position or '-'
            ws.cell(row=row, column=5).value = result.campaign.title
            ws.cell(row=row, column=6).value = float(result.overall_score) if result.overall_score else 0
            ws.cell(row=row, column=7).value = float(result.self_score) if result.self_score else 0
            ws.cell(row=row, column=8).value = float(result.supervisor_score) if result.supervisor_score else 0
            ws.cell(row=row, column=9).value = float(result.peer_score) if result.peer_score else 0
            ws.cell(row=row, column=10).value = float(result.subordinate_score) if result.subordinate_score else 0
            ws.cell(row=row, column=11).value = result.calculated_at.strftime('%d.%m.%Y') if result.calculated_at else '-'

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        excel_content = buffer.getvalue()
        buffer.close()

        response = HttpResponse(
            excel_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="custom_report_{campaign_id}.xlsx"'
        return response

    elif export_format == 'csv':
        # Generate CSV
        csv_content = generate_csv_report(results, include_fields=[
            'employee_id', 'full_name', 'department', 'position', 'campaign',
            'overall_score', 'self_score', 'supervisor_score', 'peer_score',
            'subordinate_score', 'calculated_at'
        ])

        response = HttpResponse(csv_content, content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="custom_report_{campaign_id}.csv"'
        return response

    messages.error(request, 'Etibarsız ixrac formatı.')
    return redirect('reports:custom-builder')